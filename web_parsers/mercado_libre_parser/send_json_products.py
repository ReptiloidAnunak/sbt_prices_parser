import json
import os
from pathlib import Path

import requests
import openpyxl

API_URL = os.getenv(
    "API_URL",
    "http://127.0.0.1:8010/api/products/import/"
)

XLSX_FILE = os.getenv(
    "PRODUCTS_XLSX",
    str(Path(__file__).with_name("products.xlsx"))
)

REQUIRED_COLUMNS = ["sku", "name", "price", "url", "shop"]


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_price(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        raise ValueError(f"Не удалось преобразовать цену: {value!r}")


def read_products_from_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [normalize_text(cell.value) for cell in ws[1]]
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(
            "В Excel не хватает обязательных колонок: " + ", ".join(missing)
        )

    idx = {header: headers.index(header) for header in REQUIRED_COLUMNS}
    grouped = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = normalize_text(row[idx["sku"]])
        name = normalize_text(row[idx["name"]])
        price_raw = row[idx["price"]]
        url = normalize_text(row[idx["url"]])
        shop = normalize_text(row[idx["shop"]])

        if not any([sku, name, price_raw, url, shop]):
            continue

        if not shop:
            raise ValueError(f"У строки без shop: sku={sku!r}, name={name!r}")

        product = {
            "code": sku,
            "title": name,
            "price": normalize_price(price_raw),
            "url": url,
        }

        grouped.setdefault(shop, []).append(product)

    return grouped


def send_products_grouped(file_path):
    grouped_products = read_products_from_xlsx(file_path)
    results = []

    for shop, products in grouped_products.items():
        payload = {
            "supplier": shop,
            "products": products,
            "type": "ml"
        }

        response = requests.post(API_URL, json=payload)

        result = {
            "supplier": shop,
            "count": len(products),
            "status_code": response.status_code,
            "response_text": response.text,
        }
        results.append(result)

        print(f"Supplier: {shop}")
        print(f"Products: {len(products)}")
        print(f"Status: {response.status_code}")
        print(f"Response: {response.text}")
        print("-" * 60)

    return results


if __name__ == "__main__":
    send_products_grouped(XLSX_FILE)
